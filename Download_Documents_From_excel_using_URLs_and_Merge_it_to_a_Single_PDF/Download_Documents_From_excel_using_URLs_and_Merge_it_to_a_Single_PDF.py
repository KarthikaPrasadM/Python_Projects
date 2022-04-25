# import modules
from PyQt5 import QtCore, QtGui, QtWidgets
import pandas as pd
import numpy as np
import openpyxl
import os


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(796, 244)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(10, 10, 781, 181))
        self.widget.setObjectName("widget")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.widget)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.Choose_file = QtWidgets.QLabel(self.widget)
        font = QtGui.QFont()
        font.setFamily("Arial Unicode MS")
        font.setPointSize(12)
        self.Choose_file.setFont(font)
        self.Choose_file.setTextFormat(QtCore.Qt.PlainText)
        self.Choose_file.setObjectName("Choose_file")
        self.horizontalLayout.addWidget(self.Choose_file)
        self.InputBar = QtWidgets.QLineEdit(self.widget)
        self.InputBar.setObjectName("InputBar")
        self.horizontalLayout.addWidget(self.InputBar)
        self.verticalLayout.addLayout(self.horizontalLayout)
        self.horizontalLayout_2.addLayout(self.verticalLayout)
        self.StartButton = QtWidgets.QPushButton(self.widget)
        font = QtGui.QFont()
        font.setFamily("Bahnschrift SemiBold")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.StartButton.setFont(font)
        self.StartButton.setObjectName("StartButton")
        self.horizontalLayout_2.addWidget(self.StartButton)
        self.verticalLayout_2.addLayout(self.horizontalLayout_2)
        self.label = QtWidgets.QLabel(self.widget)
        self.label.setEnabled(True)
        font = QtGui.QFont()
        font.setFamily("Segoe UI Variable Small Semibol")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.label.setAutoFillBackground(False)
        self.label.setTextFormat(QtCore.Qt.AutoText)
        self.label.setScaledContents(False)
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.verticalLayout_2.addWidget(self.label)
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 796, 26))
        self.menubar.setObjectName("menubar")
        self.menuMagellan = QtWidgets.QMenu(self.menubar)
        self.menuMagellan.setObjectName("menuMagellan")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.menubar.addAction(self.menuMagellan.menuAction())

        self.StartButton.clicked.connect(self.pressed)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.Choose_file.setText(_translate("MainWindow", "Enter file location: "))
        self.StartButton.setText(_translate("MainWindow", "START"))
        self.label.setText(_translate("MainWindow", "Enter Excel file location and click \"START\" button..."))
        self.menuMagellan.setTitle(_translate("MainWindow", "Magellan"))

    def pressed(self):
        self.label.setText("Your request is processing...")
        try:
            file_location = str(self.InputBar.text())

            path = file_location
            inputFiles = os.listdir(path)

            if "Output" in inputFiles:
                outputPath = path + "\Output"
            else:
                outputPath = path + "\Output"
                os.mkdir(outputPath)

            for file in inputFiles:
                os.chdir(path)
                if file.endswith(".xlsx"):
                    # Importing file to variable "file"
                    file_name = file.split(".")[0]

                    # Reading excel file to variable "xl" as Pandas DataFrame
                    xl = pd.ExcelFile(file)
                    # To get the number of sheets present in the excel file
                    Number_Of_Sheet_In_The_excel = len(xl.sheet_names)
                    # To get the sheet names
                    sheet_Names = xl.sheet_names
                    # print(sheet_Names)
                    ## Checking which Sheet has the data that we are looking for
                    ## Checking which sheet has the names as numbers(Ex: 1, 2, 3 etc)000
                    int_sheets = []
                    for sheet in sheet_Names:
                        sheet = sheet.replace(" ", "")
                        bool = sheet.isalpha()
                        if bool is False:
                            int_sheets.append(sheet)

                    ## Creating a new excel file to store all the output.
                    Final_file_name = file_name + " Output.xlsx"
                    writer = pd.ExcelWriter(Final_file_name, engine='openpyxl')

                    ## Filtering through each sheet to extract the data
                    for EachSheet in int_sheets:
                        os.chdir(path)

                        data = pd.read_excel(file, sheet_name=EachSheet)

                        # Drop the index and first row
                        data = data.rename(columns=data.iloc[0]).drop(data.index[0])

                        # To get the column names(Need to eliminate "NaN" values from the list)
                        Column_names = []
                        for col in data.columns:
                            if (type(col)) == str:
                                Column_names.append(col)
                            else:
                                pass

                        # To get the first column.
                        filtered_first_column = []
                        first_column = data.iloc[:, 0].tolist()
                        for q in first_column:
                            if (type(q)) == float:
                                pass
                            else:
                                filtered_first_column.append(q)

                        # Finding the position of "Write-in"
                        # If the "Write-in" is in the beginning of list drop first 4 columns from df'.
                        # If the "Write-in" is in the end of list drop last 4 columns from df'.
                        if Column_names.index("Write-in") != 0:
                            os.chdir(path)
                            # remove last element from the list
                            Filtered_Column_Names = Column_names[:-1]

                            # Replacing Header with Top Row
                            new_header = data.iloc[0]
                            data = data[1:]
                            data.columns = new_header

                            # Taking only "Total Votes" columns from df.
                            df = data["Total Votes"]

                            # Removing the column "Total Votes" related to "Write-in"
                            # ex: Removing the last column of a dataframe
                            df = df.iloc[:, :-1]

                            # Replace the column names(header) with a list.
                            df = df.set_axis(Filtered_Column_Names, axis=1, inplace=False)

                            # adding the list to the dataframe as column
                            df[' '] = filtered_first_column

                            # Moving the added column to the frond of the df
                            df = df[[' '] + [col for col in df.columns if col != ' ']]

                            # Drop last row
                            # by selecting all rows except last row
                            df = df.iloc[:-1, :]

                            # Save as xlsx and skip the index
                            # df.to_excel(EachSheet + "output.xlsx", index=False)
                            # print(df)
                            df.to_excel(writer, EachSheet, index=False)
                            os.chdir(outputPath)
                            writer.save()

                        else:
                            # set path
                            os.chdir(outputPath)

                            # Removing first element from the list
                            Filtered_Column_Names = Column_names[1:]

                            # Replacing Header with Top Row
                            new_header = data.iloc[0]
                            data = data[1:]
                            data.columns = new_header

                            # Taking only "Total Votes" columns from df.
                            df = data["Total Votes"]

                            # Removing the column "Total Votes" related to "Write-in"
                            # ex: Removing the first column of a dataframe
                            # Drop first column of dataframe
                            df = df.iloc[:, 1:]

                            # Replace the column names(header) with a list.
                            df = df.set_axis(Filtered_Column_Names, axis=1, inplace=False)

                            # adding the list to the dataframe as column
                            df[' '] = filtered_first_column

                            # Moving the added column to the frond of the df
                            df = df[[' '] + [col for col in df.columns if col != ' ']]

                            # Drop last row
                            # by selecting all rows except last row
                            df = df.iloc[:-1, :]

                            # Save as xlsx and skip the index
                            # df.to_excel(EachSheet + "output.xlsx", index=False)
                            # print(df)
                            df.to_excel(writer, EachSheet, index=False)
                            writer.save()

            # create an Empty DataFrame object
            dfFinal = pd.DataFrame()

            for file1 in inputFiles:
                os.chdir(path)
                if file1.endswith(".xlsx"):
                    print(file1)
                    ### STAGE_1
                    # Importing file to variable "file"
                    file2 = file1.split(".")[0]
                    outputXLSX = file2 + ' Output.xlsx'
                    # Reading excel file as df using Pandas
                    df = pd.read_excel(file1, sheet_name="Registered Voters")

                    # To get the first column as list and removing 'Total:' from the list
                    first_column = df.iloc[:, 0].tolist()
                    first_column.remove('Total:')
                    # print("first_column")
                    # print(first_column)
                    ### STAGE_2
                    ## 1. Filter through each sheet.
                    ## 2. Spacify sheet name
                    wb1 = openpyxl.load_workbook(outputXLSX)
                    sheetNames = (wb1.sheetnames)

                    finaldf = pd.DataFrame()

                    for sheetnames in sheetNames:
                        # Reading excel file as df using Pandas by name
                        os.chdir(path)
                        out_df = pd.read_excel(outputXLSX, sheet_name=sheetnames)

                        # To get the first column as list and removing 'Total:' from the list
                        out_first_column = out_df.iloc[:, 0].tolist()
                        # print("out_first_column")
                        # print(out_first_column)
                        ### Checking how many rows are in the Output file and arrange it
                        # Create an empty dataframe and add the values into it
                        # extract the header from the "out_file" df
                        New_Column_Names = list(out_df.columns.values)
                        New_Column_Names[0] = "Key"
                        # Create a new df and add column names from a list
                        df1 = pd.DataFrame(columns=New_Column_Names)
                        # print(df1)
                        # Taking the dataframe as a list of list
                        listOfDFRows = out_df.to_numpy().tolist()
                        # print(listOfDFRows)
                        Check_List = []
                        count = 0
                        for i in first_column:
                            if i in out_first_column:
                                # Convert a dataframe to the list of rows i.e. list of lists
                                itemm = listOfDFRows[count]
                                df1.loc[len(df1)] = itemm
                                # print(count)
                                # print(itemm)
                                count = count + 1
                            else:
                                lenOfNames = len(New_Column_Names)
                                spaces = list(" " * lenOfNames)
                                df1.loc[len(df1)] = spaces

                        # print(df1)
                        finaldf = pd.concat([finaldf, df1], axis=1)
                    del finaldf['Key']
                    idx = 0
                    finaldf.insert(loc=idx, column=' ', value=first_column)
                    os.chdir(outputPath)
                    finaldf.to_excel(file2 + " Final Out.xlsx", index=False)
                    # finaldf.drop(finaldf.index, inplace=True)
            self.label.setText("Completed!")

        except:
            pass
if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())